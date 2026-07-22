from __future__ import annotations

import hashlib
import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from app.rag.models import KnowledgeChunk, KnowledgeDocument


class KnowledgeDocumentError(ValueError):
    pass


class MarkdownKnowledgeLoader:
    required_metadata = {
        "title",
        "maintainer",
        "updated_at",
        "domain",
        "visibility",
        "status",
        "source_refs",
    }
    allowed_visibilities = {"public", "user", "admin"}

    def __init__(self, knowledge_dir: Path, repository_root: Path) -> None:
        self.knowledge_dir = knowledge_dir.resolve()
        self.repository_root = repository_root.resolve()

    def load(self) -> list[KnowledgeDocument]:
        if not self.knowledge_dir.is_dir():
            raise KnowledgeDocumentError(f"Knowledge directory does not exist: {self.knowledge_dir}")

        documents: list[KnowledgeDocument] = []
        for path in sorted(self.knowledge_dir.rglob("*.md")):
            if path.name.lower() == "readme.md":
                continue
            document = self._load_markdown(path)
            if document is not None:
                documents.append(document)
        documents.extend(self._load_structured_sources())
        return documents

    def _load_structured_sources(self) -> list[KnowledgeDocument]:
        manifest_path = self.knowledge_dir / "structured-sources.json"
        if not manifest_path.exists():
            return []
        payload = json.loads(manifest_path.read_text(encoding="utf-8"))
        documents: list[KnowledgeDocument] = []
        for entry in payload.get("sources", []):
            if entry.get("status") != "approved":
                continue
            missing = self.required_metadata - entry.keys()
            if missing:
                raise KnowledgeDocumentError(
                    f"{manifest_path}: source {entry.get('path')} missing metadata {sorted(missing)}"
                )
            path = (self.knowledge_dir / str(entry["path"])).resolve()
            if self.knowledge_dir not in path.parents or not path.is_file():
                raise KnowledgeDocumentError(f"Invalid structured source path: {path}")
            if entry["visibility"] not in self.allowed_visibilities:
                raise KnowledgeDocumentError(f"{path}: invalid visibility")
            content = self._extract_structured(path)
            if not content.strip():
                raise KnowledgeDocumentError(f"{path}: empty document")
            source = path.relative_to(self.repository_root).as_posix()
            documents.append(KnowledgeDocument(
                source=source,
                title=str(entry["title"]),
                domain=str(entry["domain"]),
                visibility=str(entry["visibility"]),
                updated_at=str(entry["updated_at"]),
                content_hash=hashlib.sha256(path.read_bytes()).hexdigest(),
                content=content,
            ))
        return documents

    def _extract_structured(self, path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix == ".pdf":
            return "\n\n".join(
                f"# 第 {index} 页\n{page.extract_text() or ''}"
                for index, page in enumerate(PdfReader(str(path)).pages, 1)
            )
        if suffix == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            sections: list[str] = []
            try:
                for sheet in workbook.worksheets:
                    rows = [
                        " | ".join("" if value is None else str(value) for value in row)
                        for row in sheet.iter_rows(values_only=True)
                    ]
                    sections.append(f"# 工作表 {sheet.title}\n" + "\n".join(rows))
            finally:
                workbook.close()
            return "\n\n".join(sections)
        if suffix == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
            return "\n\n".join(
                f"# 数据行 {index}\n" + "\n".join(f"{key}: {value}" for key, value in row.items())
                for index, row in enumerate(rows, 1)
            )
        if suffix == ".json":
            value = json.loads(path.read_text(encoding="utf-8"))
            return "# JSON 结构化规则\n" + json.dumps(value, ensure_ascii=False, indent=2)
        raise KnowledgeDocumentError(f"Unsupported structured source: {path.suffix}")

    def split(self, document: KnowledgeDocument, *, max_chars: int = 320) -> list[KnowledgeChunk]:
        sections = self._sections(document.content, document.title)
        packed: list[str] = []
        current = ""
        for section in sections:
            if current and len(current) + len(section) + 2 > max_chars:
                packed.append(current.strip())
                current = section
            else:
                current = f"{current}\n\n{section}".strip()
        if current:
            packed.append(current.strip())

        bounded: list[str] = []
        for content in packed:
            if len(content) <= max_chars:
                bounded.append(content)
                continue
            for offset in range(0, len(content), max_chars):
                bounded.append(content[offset : offset + max_chars].strip())
        packed = [content for content in bounded if content]

        chunks: list[KnowledgeChunk] = []
        for index, content in enumerate(packed):
            chunk_id = hashlib.sha256(
                f"{document.source}\0{document.content_hash}\0{index}\0{content}".encode("utf-8")
            ).hexdigest()
            chunks.append(
                KnowledgeChunk(
                    id=chunk_id,
                    source=document.source,
                    title=document.title,
                    domain=document.domain,
                    visibility=document.visibility,
                    updated_at=document.updated_at,
                    content_hash=document.content_hash,
                    chunk_index=index,
                    content=content,
                )
            )
        return chunks

    def _load_markdown(self, path: Path) -> KnowledgeDocument | None:
        raw = path.read_text(encoding="utf-8")
        metadata, content = self._parse_front_matter(raw, path)
        missing = self.required_metadata - metadata.keys()
        if missing:
            raise KnowledgeDocumentError(f"{path}: missing metadata {sorted(missing)}")
        if metadata["status"] != "approved":
            return None
        if metadata["visibility"] not in self.allowed_visibilities:
            raise KnowledgeDocumentError(f"{path}: invalid visibility")
        if not content.strip():
            raise KnowledgeDocumentError(f"{path}: empty document")

        source = path.resolve().relative_to(self.repository_root).as_posix()
        content_hash = hashlib.sha256(raw.encode("utf-8")).hexdigest()
        return KnowledgeDocument(
            source=source,
            title=metadata["title"],
            domain=metadata["domain"],
            visibility=metadata["visibility"],
            updated_at=metadata["updated_at"],
            content_hash=content_hash,
            content=content.strip(),
        )

    def _parse_front_matter(self, raw: str, path: Path) -> tuple[dict[str, str], str]:
        normalized = raw.lstrip("\ufeff").replace("\r\n", "\n")
        if not normalized.startswith("---\n"):
            raise KnowledgeDocumentError(f"{path}: YAML front matter is required")
        parts = normalized.split("---\n", 2)
        if len(parts) != 3:
            raise KnowledgeDocumentError(f"{path}: invalid YAML front matter")

        metadata: dict[str, str] = {}
        for line in parts[1].splitlines():
            if not line or line.startswith((" ", "-")) or ":" not in line:
                continue
            key, value = line.split(":", 1)
            metadata[key.strip()] = value.strip().strip('"\'')
        return metadata, parts[2]

    def _sections(self, content: str, title: str) -> list[str]:
        sections: list[str] = []
        heading = f"# {title}"
        body: list[str] = []
        for line in content.splitlines():
            if re.match(r"^#{1,3}\s+", line):
                if body:
                    sections.append(f"{heading}\n" + "\n".join(body).strip())
                    body = []
                heading = line.strip()
            else:
                body.append(line)
        if body:
            sections.append(f"{heading}\n" + "\n".join(body).strip())
        return [section for section in sections if section.strip()]
